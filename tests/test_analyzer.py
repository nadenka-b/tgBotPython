import pytest
from io import BytesIO


@pytest.fixture
def mock_db():
    """Mock объект БД для тестирования анализатора"""
    class MockDB:
        pass
    return MockDB()


class TestAnalyzer:
    """Тесты для анализатора данных"""

    def test_analyzer_excel_workbook_creation(self, mock_db):
        """Проверка создания Excel workbook"""
        from analyzer.analyzer import DataAnalyzer
        import openpyxl

        analyzer = DataAnalyzer(mock_db)
        wb = analyzer._create_excel_workbook()

        assert wb is not None
        assert isinstance(wb, openpyxl.Workbook)
        assert wb.active is not None

    def test_analyzer_bytes_io_conversion(self, mock_db):
        """Проверка преобразования Excel в BytesIO"""
        from analyzer import DataAnalyzer

        analyzer = DataAnalyzer(mock_db)
        wb = analyzer._create_excel_workbook()
        wb.active['A1'] = "Тестовые данные"

        buffer = analyzer._get_bytes_io(wb)

        assert isinstance(buffer, BytesIO)
        assert buffer.tell() == 0
        assert len(buffer.getvalue()) > 0

    def test_analyzer_buffer_seekable(self, mock_db):
        """Проверка что буфер можно читать"""
        from analyzer import DataAnalyzer

        analyzer = DataAnalyzer(mock_db)
        wb = analyzer._create_excel_workbook()
        wb.active['A1'] = "Данные"

        buffer = analyzer._get_bytes_io(wb)

        data = buffer.read()
        assert len(data) > 0

        buffer.seek(0)
        assert buffer.tell() == 0

    def test_analyzer_column_auto_adjust_min_width(self, mock_db):
        """Проверка минимальной ширины при подстройке"""
        from analyzer import DataAnalyzer

        analyzer = DataAnalyzer(mock_db)
        wb = analyzer._create_excel_workbook()
        ws = wb.active

        ws['A1'] = "abc"
        analyzer._auto_adjust_column(ws, 'A')

        width = ws.column_dimensions['A'].width
        assert width >= 10  # Минимальная ширина 10

    def test_analyzer_column_auto_adjust_long_text(self, mock_db):
        """Проверка подстройки для длинного текста"""
        from analyzer import DataAnalyzer

        analyzer = DataAnalyzer(mock_db)
        wb = analyzer._create_excel_workbook()
        ws = wb.active

        long_text = "Это очень длинный текст для проверки автоматической подстройки ширины колонки"
        ws['A1'] = long_text

        analyzer._auto_adjust_column(ws, 'A')

        width = ws.column_dimensions['A'].width
        assert width > 10

    def test_analyzer_multiple_rows_adjust(self, mock_db):
        """Проверка подстройки для нескольких строк"""
        from analyzer import DataAnalyzer

        analyzer = DataAnalyzer(mock_db)
        wb = analyzer._create_excel_workbook()
        ws = wb.active

        ws['A1'] = "Короткий"
        ws['A2'] = "Это намного более длинный текст в ячейке"
        ws['A3'] = "abc"

        analyzer._auto_adjust_column(ws, 'A')

        width = ws.column_dimensions['A'].width
        assert width > 10

    def test_analyzer_initialization_with_db(self, mock_db):
        """Проверка инициализации анализатора с БД"""
        from analyzer import DataAnalyzer

        analyzer = DataAnalyzer(mock_db)

        assert analyzer.db == mock_db
        assert analyzer is not None
