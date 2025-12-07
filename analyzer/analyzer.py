import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from sqlalchemy import and_, case, func, or_
from sqlalchemy.orm import Session
from typing import Union

from database import Database, FilterCombination, Statistics

logger = logging.getLogger(__name__)


class DataAnalyzer:
    """Анализатор данных поступления в КФУ"""

    def __init__(self, db: Database):
        self.db = db

    def _create_excel_workbook(self) -> openpyxl.Workbook:
        """Создает новый Excel файл"""
        return openpyxl.Workbook()

    def _style_header(self, ws, row_num):
        """Стилизует заголовочную строку"""
        for cell in ws[row_num]:
            if cell.value is not None:
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)

    def _get_bytes_io(self, wb: openpyxl.Workbook) -> BytesIO:
        """Преобразует Excel в BytesIO (для отправки)"""
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _auto_adjust_column(self, ws, column_letter):
        column = ws[column_letter]
        max_length = 0

        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass

        adjusted_width = max(10, max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    def _analyze_categoties(self, session: Session, filter_id, filter_note) -> tuple[dict[str, dict[str, int]], int]:
        category_stats = session.query(
            Statistics.admission_category,
            Statistics.available_places,
            func.count(Statistics.id).label('total_count'),
            func.count(
                case(
                    (and_(
                        Statistics.agreement == 'да',
                        filter_note
                    ), 1)
                )
            ).label('agreed_count')
        ).filter(
            filter_id
        ).group_by(
            Statistics.admission_category,
            Statistics.available_places
        ).all()

        admitted_by_category = {}
        occupied_places = 0
        for stat_row in category_stats:
            cat_name = stat_row.admission_category
            if not cat_name or cat_name == 'общий конкурс':
                continue
            agreed_count = stat_row.agreed_count or 0
            available_places = stat_row.available_places or 0
            if cat_name == 'без вступительных испытаний':
                available_places = agreed_count
            admitted_by_category[cat_name] = {
                'available_places': available_places,
                'agreed_count': agreed_count,
                'total': min(available_places, agreed_count)
            }
            occupied_places += admitted_by_category[cat_name]['total']
            logger.debug(
                f"✅ Категория '{cat_name}': согласились {agreed_count}, мест {available_places}, заняли {admitted_by_category[cat_name]['total']}")
        return admitted_by_category, occupied_places

    def analyze_speciality(self, filters: dict) -> Union[BytesIO, dict]:
        """Анализ конкретного направления"""
        session = self.db.get_session()
        try:
            combo = session.query(FilterCombination).filter(
                FilterCombination.level_value == filters['level'],
                FilterCombination.inst_value == filters['inst'],
                FilterCombination.faculty_value == filters['faculty'],
                FilterCombination.speciality_value == filters['speciality'],
                FilterCombination.typeofstudy_value == filters['typeofstudy'],
                FilterCombination.category_value == filters['category'],
            ).first()

            if not combo:
                logger.warning("⚠️ Комбинация не найдена")
                return {'error': '❌ Не существует такого сочетания параметров'}

            filter_id = Statistics.filter_combination_id == combo.id
            filter_note = or_(Statistics.note.is_(None),
                              ~Statistics.note.contains("не сданы один или несколько экзаменов"))

            total_apps = session.query(func.count(Statistics.id)).filter(
                filter_id
            ).scalar() or 0

            if not total_apps:
                logger.warning(f"⚠️ Нет статистики для комбинации {combo.id}")
                return {'error': '❌ Не существует данных для такого сочетания параметров'}

            total_participants = session.query(func.count(Statistics.id)).filter(
                filter_id, filter_note).scalar() or 0

            total_available_places = session.query(Statistics.available_places).filter(
                filter_id,
                Statistics.admission_category == 'общий конкурс',
            ).limit(1).scalar()

            if not total_available_places:
                logger.warning(
                    "Не найдены доступные места по этому направлению")
                return {'error': '❌ Не найдены доступные места по этому направлению'}

            admitted_by_category, occupied_places = self._analyze_categoties(
                session, filter_id, filter_note)

            remaining_places = total_available_places - occupied_places

            logger.debug(
                f"Всего мест: {total_available_places}, занято: {occupied_places}, осталось: {remaining_places}")

            general_participants = session.query(Statistics).filter(
                filter_id, filter_note,
                Statistics.admission_category == 'общий конкурс',
                Statistics.agreement == 'да',
            ).order_by(Statistics.score.desc()).limit(remaining_places).all()

            scores: list = [
                s.score for s in general_participants if s.score is not None]

            wb = self._create_excel_workbook()
            ws = wb.active
            if not ws:
                return {'error': 'Ошибка создания файла'}
            ws.title = "Анализ направления"

            ws['A1'] = "АНАЛИЗ НАПРАВЛЕНИЯ"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')

            ws['A2'] = "Название специальности:"
            ws['B2'] = str(combo.speciality_name)

            ws['A3'] = "Всего заявлений:"
            ws['B3'] = total_apps

            ws['A4'] = "Всего конкурсантов:"
            ws['B4'] = total_participants

            ws['A5'] = "Всего мест:"
            ws['B5'] = total_available_places

            ws['A6'] = "Мест на общий конкурс:"
            ws['B6'] = remaining_places

            ws['A7'] = "Заявлений на место:"
            ws['B7'] = round(total_participants / remaining_places,
                             2) if remaining_places > 0 else 0

            ws['A8'] = "Средний балл:"
            ws['B8'] = round(sum(scores) / len(scores), 1) if scores else None

            ws['A9'] = "Минимальный балл:"
            ws['B9'] = min(scores) if scores else None

            ws['A10'] = "Максимальный балл:"
            ws['B10'] = max(scores) if scores else None

            if admitted_by_category:
                ws['A12'] = "Специальные категории"
                ws['A12'].font = Font(bold=True, size=12)

                headers: list = ['Категория', 'Доступно мест',
                                 'Согласились', 'Занято мест']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=13, column=col_num)
                    cell.value = header

                self._style_header(ws, 13)
                row = 14
                for cat_name, cat_data in admitted_by_category.items():
                    ws.cell(row=row, column=1).value = cat_name
                    ws.cell(
                        row=row, column=2).value = cat_data['available_places']
                    ws.cell(row=row, column=3).value = cat_data['agreed_count']
                    ws.cell(row=row, column=4).value = cat_data['total']
                    row += 1

            self._auto_adjust_column(ws, 'A')
            for col in ['B', 'C', 'D']:
                ws.column_dimensions[col].width = 15

            logger.info(
                f"✅ Анализ направления {combo.speciality_name}: {total_apps} заявлений")
            return self._get_bytes_io(wb)

        except Exception as e:
            logger.error(f"❌ Ошибка анализа направления: {e}")
            return {'error': "❌ Ошибка анализа направления"}
        finally:
            session.close()

    def analyze_institute(self, filters: dict) -> Union[BytesIO, dict]:
        """Анализ популярности направлений в институте"""
        session = self.db.get_session()
        try:
            base_filter = [
                FilterCombination.level_value == filters.get('level'),
                FilterCombination.inst_value == filters.get('inst'),
                FilterCombination.faculty_value == filters.get('faculty'),
                FilterCombination.category_value == filters.get('category'),
            ]
            combos = session.query(FilterCombination).filter(
                *base_filter,).all()

            if not combos:
                logger.warning(
                    f"⚠️ Нет такого сочетания параметров для института {filters['inst']}")
                return {
                    'error': '❌ Не существует данных для такого сочетания параметров'
                }

            faculty_name = combos[0].faculty_name
            combo_ids = [c.id for c in combos]

            filter_id = Statistics.filter_combination_id.in_(combo_ids)
            filter_applicant_id = or_(Statistics.epgu_id.isnot(None),
                                      Statistics.applicant_id.isnot(None))
            filter_note = or_(~Statistics.note.contains(
                "не сданы один или несколько экзаменов"),
                Statistics.note.is_(None))

            unique_applicant_count = session.query(func.count(Statistics.id.distinct())).filter(
                filter_id, filter_applicant_id).scalar() or 0

            unique_participants_count = session.query(func.count(Statistics.id.distinct())).filter(
                filter_id, filter_applicant_id, filter_note).scalar() or 0

            logger.debug(
                f"Уникальных заявлений: {unique_applicant_count}, уникальных конкурсантов: {unique_participants_count}")

            spec_category_all = session.query(
                FilterCombination.speciality_name,
                FilterCombination.speciality_value,
                Statistics.available_places,
                Statistics.admission_category,
                func.count(Statistics.id).label('total_apps'),
            ).join(
                Statistics, Statistics.filter_combination_id == FilterCombination.id
            ).filter(filter_id, filter_note,
                     ).group_by(
                FilterCombination.speciality_value,
                FilterCombination.speciality_name,
                Statistics.available_places,
                Statistics.admission_category,
            ).all()

            spec_category_admitted = session.query(
                FilterCombination.speciality_name,
                FilterCombination.speciality_value,
                Statistics.admission_category,
                func.count(Statistics.id).label('admitted_count'),
            ).join(
                Statistics, Statistics.filter_combination_id == FilterCombination.id
            ).filter(
                *base_filter,
                filter_note,
                Statistics.agreement == 'да',
            ).group_by(
                FilterCombination.speciality_value,
                FilterCombination.speciality_name,
                Statistics.admission_category,
            ).all()

            if not spec_category_all:
                logger.warning(f"⚠️ Нет статистики для специальностей")
                return {'error': '❌ Нет статистики для специальностей'}

            specialities_map = {}
            for row in spec_category_all:
                spec_key = row.speciality_name
                if spec_key not in specialities_map:
                    specialities_map[spec_key] = {
                        'name': spec_key,
                        'total_places': 0,
                        'categories': {},
                        'total_apps': 0,
                    }
                cat = row.admission_category or 'общий конкурс'
                total_apps = row.total_apps or 0
                specialities_map[spec_key]['total_apps'] += total_apps
                max_places_for_cat = row.available_places or 0
                specialities_map[spec_key]['categories'][cat] = {
                    'total_applications': total_apps,
                    'total_participants': 0,
                    'max_places': max_places_for_cat,
                    'admitted': 0,
                }

            for row in spec_category_admitted:
                spec_key = row.speciality_name
                cat = row.admission_category or 'общий конкурс'
                if spec_key in specialities_map and cat in specialities_map[spec_key]['categories']:
                    specialities_map[spec_key]['categories'][cat]['admitted'] = row.admitted_count or 0
                    if cat == 'без вступительных испытаний':
                        specialities_map[spec_key]['categories'][cat]['max_places'] = specialities_map[spec_key]['categories'][cat]['admitted']

            specialities_data = []
            for spec_key, spec_info in specialities_map.items():
                total_places = spec_info['categories'].get(
                    'общий конкурс', {}).get('max_places', 0) or 0
                specialities_map[spec_key]['total_places'] = total_places

                occupied_by_special = 0
                for cat, cat_info in spec_info['categories'].items():
                    if cat == 'общий конкурс':
                        continue
                    max_places = cat_info.get('max_places', 0)
                    admitted_in_cat = cat_info.get('admitted', 0)
                    occupied = min(admitted_in_cat, max_places)
                    occupied_by_special += occupied

                remaining_places = total_places - occupied_by_special
                total_apps = spec_info['total_apps']
                applicants_per_place = round(
                    total_apps / total_places, 2) if total_places > 0 else 0
                specialities_data.append({
                    'name': spec_info['name'],  # название специальности
                    'total_applications': total_apps,  # всего заявлений
                    'total_places': total_places,  # всего мест
                    # занято мест специальными категориями
                    'occupied_by_special_categories': occupied_by_special,
                    'remaining_places': remaining_places,  # осталось мест на общий конкурс
                    'applicants_per_place': applicants_per_place,  # заявлений на одно место
                })

            specialities_data.sort(
                key=lambda x: x['applicants_per_place'], reverse=True)

            wb = self._create_excel_workbook()
            ws = wb.active
            if not ws:
                return {'error': 'Ошибка создания файла'}
            ws.title = "Анализ института"

            ws['A1'] = "АНАЛИЗ ИНСТИТУТА"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:F1')

            ws['A2'] = f"Название: {faculty_name}"
            ws['A3'] = f"Уникальных заявителей: {unique_applicant_count}"
            ws['A4'] = f"Уникальных конкурсантов: {unique_participants_count}"
            ws['A5'] = f"Специальностей: {len(specialities_data)}"

            headers: list = ['Специальность', 'Ранг', 'Заявления',
                             'Мест', 'Спец. кат.', 'Осталось', 'Конкурс']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col_num)
                cell.value = header

            self._style_header(ws, 7)

            row = 8
            for idx, spec in enumerate(specialities_data, 1):
                ws.cell(row=row, column=1).value = spec['name']
                ws.cell(row=row, column=2).value = idx
                ws.cell(row=row, column=3).value = spec['total_applications']
                ws.cell(row=row, column=4).value = spec['total_places']
                ws.cell(
                    row=row, column=5).value = spec['occupied_by_special_categories']
                ws.cell(row=row, column=6).value = spec['remaining_places']
                ws.cell(row=row, column=7).value = spec['applicants_per_place']
                row += 1

            for col in ['B', 'C', 'D', 'E', 'F', 'G']:
                ws.column_dimensions[col].width = 12
            self._auto_adjust_column(ws, 'A')

            logger.info(
                f"✅ Анализ института {faculty_name}: {len(specialities_data)} специальностей")

            return self._get_bytes_io(wb)
        except Exception as e:
            logger.error(f"❌ Ошибка анализа института: {e}")
            return {'error': '❌ Не удалось проанализировать институт'}
        finally:
            session.close()

    def analyze_university(self, filters: dict) -> Union[BytesIO, dict]:
        """Анализ всего университета"""
        session = self.db.get_session()
        try:
            combos = session.query(FilterCombination).filter(
                FilterCombination.level_value == filters.get('level'),
                FilterCombination.inst_value == filters.get('inst'),
                FilterCombination.category_value == filters.get('category'),
            ).all()

            if not combos:
                logger.warning(
                    f"⚠️ Нет такого сочетания параметров для университета {filters['inst']}")
                return {
                    'error': '❌ Не существует данных для такого сочетания параметров'
                }

            inst_name = combos[0].inst_name
            combo_ids = [c.id for c in combos]

            filter_id = Statistics.filter_combination_id.in_(combo_ids)
            filter_applicant_id = or_(Statistics.epgu_id.isnot(None),
                                      Statistics.applicant_id.isnot(None))
            filter_note = or_(~Statistics.note.contains(
                "не сданы один или несколько экзаменов"),
                Statistics.note.is_(None))

            unique_applicant_count = session.query(func.count(Statistics.id.distinct())).filter(
                filter_id, filter_applicant_id).scalar() or 0

            unique_participants_count = session.query(func.count(Statistics.id.distinct())).filter(
                filter_id, filter_applicant_id, filter_note).scalar() or 0

            logger.debug(
                f"Университет {inst_name}: {unique_applicant_count} заявок, {unique_participants_count} конкурсантов")

            faculties_all_stats = session.query(
                FilterCombination.faculty_value,
                FilterCombination.faculty_name,
                func.count(Statistics.id.distinct()).label('total_apps')
            ).join(
                Statistics, Statistics.filter_combination_id == FilterCombination.id
            ).filter(
                filter_id,
                filter_applicant_id,
            ).group_by(
                FilterCombination.faculty_value,
                FilterCombination.faculty_name,
            ).all()

            faculties_unique_stats = session.query(
                FilterCombination.faculty_value,
                FilterCombination.faculty_name,
                func.count(Statistics.id.distinct()).label(
                    'unique_participants'),

            ).join(
                Statistics, Statistics.filter_combination_id == FilterCombination.id
            ).filter(
                filter_id,
                filter_applicant_id,
                filter_note,
            ).group_by(
                FilterCombination.faculty_value,
                FilterCombination.faculty_name,
            ).all()

            if not faculties_all_stats:
                logger.warning(
                    f"⚠️ Нет статистики для институтов по университету {inst_name}")
                return {
                    'total_unique_applicants': unique_applicant_count,
                    'total_unique_participants': unique_participants_count,
                    'faculties': [],
                    'most_popular': None,
                    'least_popular': None,
                }

            faculties_data = {}
            for faculty in faculties_all_stats:
                faculties_data[faculty.faculty_name] = {
                    'unique_applications': faculty.total_apps or 0,
                    'unique_participants': 0
                }

            for faculty in faculties_unique_stats:
                if faculty.faculty_name in faculties_data:
                    faculties_data[faculty.faculty_name]['unique_participants'] = faculty.unique_participants or 0

            faculty_list = []
            for faculty_name, data in faculties_data.items():
                if data['unique_applications'] > 0:
                    faculty_list.append({
                        'name': faculty_name,
                        'unique_applications': data['unique_applications'],
                        'unique_participants': data['unique_participants'],
                    })

            faculty_list.sort(
                key=lambda x: x['unique_participants'], reverse=True)

            wb = self._create_excel_workbook()
            ws = wb.active
            if not ws:
                return {}
            ws.title = "Анализ ВУЗа"

            ws['A1'] = "АНАЛИЗ УНИВЕРСИТЕТА"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')

            ws['A2'] = f"Университет: {inst_name}"
            ws['A3'] = f"Уникальных заявителей: {unique_applicant_count}"
            ws['A4'] = f"Уникальных конкурсантов: {unique_participants_count}"
            ws['A5'] = f"Факультетов: {len(faculty_list)}"

            headers: list = ['Факультет', 'Ранг', 'Заявления', 'Конкурс']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col_num)
                cell.value = header

            self._style_header(ws, 7)

            row = 8
            for idx, fac in enumerate(faculty_list, 1):
                ws.cell(row=row, column=1).value = fac['name']
                ws.cell(row=row, column=2).value = idx
                ws.cell(row=row, column=3).value = fac['unique_applications']
                ws.cell(row=row, column=4).value = fac['unique_participants']
                row += 1

            for col in ['B', 'C', 'D']:
                ws.column_dimensions[col].width = 12
            self._auto_adjust_column(ws, 'A')

            logger.info(
                f"✅ Анализ университета {inst_name}: {len(faculty_list)} институтов"
            )
            return self._get_bytes_io(wb)

        except Exception as e:
            logger.error(f"❌ Ошибка анализа университета: {e}")
            return {'error': '❌ Не удалось проанализировать университет'}
        finally:
            session.close()
